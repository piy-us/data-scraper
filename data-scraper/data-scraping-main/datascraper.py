import requests
from bs4 import BeautifulSoup
import openpyxl

from difflib import SequenceMatcher

#checking whether the name is matching
def matching(a, b):
    return SequenceMatcher(None, a, b).ratio()

def add_search(college):
    #requestheader copied from the browser
    address=' '
    header={
        'user-agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/105.0.0.0 Safari/537.36',
        'accept':'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
        'accept-encoding':'gzip, deflate',
        'accept-language':'en-US,en;q=0.9'
        }

    #search url
    searchword=college.split(' ')
    URL="https://www.indcareer.com/find/all-colleges?qqq="+'+'.join(searchword)+'&submit=Search'
    print(URL)

    responses = requests.get(URL,headers=header)
    
    html=BeautifulSoup(responses.content,"html.parser")
    
    if responses.status_code==200:                               # check if response is successful
        #selct div element which contains data      
        maindiv=html.find('div',attrs={'class':'media'})
        if maindiv!=None:                                          #check if there's info about college
          if matching(maindiv.find('a').string,college)>=0.5:
              collegesite=maindiv.find('a')['href']
              #send request to college specific webpage                 
              responses = requests.get('https://www.indcareer.com/'+collegesite,headers=header,stream=True)

              if responses.status_code==200:
                  soup=BeautifulSoup(responses.content,"html.parser")
                  
                  tbody=soup.find('table')
                  rows=tbody.find_all('tr')
                  for i in rows:
                      try:
                          if i.find('th').string.strip()== 'Address':
                              td=i.find('td')
                              if td.string==None:
                                  address=str(td).replace("<br>", " ").strip('<td>')
                                  return(address.replace("<br/>", " "))
                              else:
                                  address=td.string.strip()
                                  return address
                      except:
                          continue        
    return address
    pass

# college_list is our excel file
wb=openpyxl.load_workbook('college_list.xlsx')
sheet_obj = wb.active
m_row = 502             #max rows
print(m_row)
for i in range(2, m_row + 1):
    print(i)
    cell_obj = sheet_obj.cell(row = i, column = 2)
    address=add_search(cell_obj.value)
    sheet_obj.cell(row=i,column=4).value=address

wb.save('college_list.xlsx')        #save file